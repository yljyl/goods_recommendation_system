import os
import re

from django.core.wsgi import get_wsgi_application
from openpyxl import load_workbook
from user_management.utils.configread import config_read
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "goods_recommendation_system.settings")
application = get_wsgi_application()
import django
django.setup()
# import emoji

# 数据读取，数据清理，并写入DB
# 获取files目录下的所有.xlsx文件名
from core_recommendations.models import Category, Goods


# dir_path = 'files/'
dir_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')

files = [f for f in os.listdir(dir_path) if f.endswith('.xlsx')]

# 存储文件名的变量
category = []

# 遍历文件
for file in files:
    file_path = os.path.join(dir_path, file)

    # 打开Excel文件
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 获取标题行的内容
    headers = [cell.value for cell in sheet[1]]

    name = file[:-5]
    print(name)
    # 保存分类
    category, created = Category.objects.get_or_create(
        name=name
    )
    if created:
        print("插入分类成功！")
    else:
        print('标签已存在')
    # 判断标题行的内容是否符合要求
    expected_headers = ['产品名称', '图片地址',  '价格', '详情', '销售量', '省份', '城市', '商家', '链接']
    if headers == expected_headers:

        if category is not None:
            # 遍历每一行内容
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # 存储每行内容的变量
                # 分别存储每列的数据
                name, img, price ,content, sales, province, city, business, links = row

                '''
                print('名称:', name)
                print('简介:', intro)
                print('项目图片:', img)
                print('分类标签:', tag_name)
                print('链接:', link)
                print('人均价格:', price)
                print('地址:', address)
                '''

                # 数据清洗
                #intro = emoji.demojize(intro.replace("\r", ""))
                good_img = 'http://localhost:8000/media/goods/'+img

                # 保存数据
                try:
                    Goods.objects.create(
                        name=name,
                        category_id=category.id,
                        img=good_img,
                        price=price,
                        content=content,
                        sales=sales,
                        province=province,
                        city=city,
                        business=business,
                        links=links,
                        views=0
                    )
                    print("插入数据成功！")
                except Exception as e:
                    pass

    else:
        print(f'标题行不符合要求: {file}')

    # 关闭Excel文件
    workbook.close()
